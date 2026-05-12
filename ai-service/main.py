import os
import re
import uuid
import zipfile
from typing import Any, Dict, List

from fastapi import FastAPI, HTTPException
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

app = FastAPI()
TEMP_DIR = "/app/data/exports/temp"

class MappingItem(BaseModel):
    fieldId: str = Field(alias="field_id")
    cell: str

class SheetMapping(BaseModel):
    sheet: str
    mappings: List[MappingItem]

class SingleExportRequest(BaseModel):
    templatePath: str = Field(alias="template_path")
    sheet: str = Field(alias="sheet", default="Sheet1")
    mappings: List[MappingItem] = Field(alias="mappings")
    recordData: Dict[str, Any] = Field(alias="record_data")
    outputName: str = Field(alias="output_name")

class BatchRecord(BaseModel):
    data: Dict[str, Any]
    outputName: str = Field(alias="output_name")

class BatchExportRequest(BaseModel):
    templatePath: str = Field(alias="template_path")
    sheet: str = Field(alias="sheet", default="Sheet1")
    mappings: List[MappingItem] = Field(alias="mappings")
    records: List[BatchRecord]
    outputZipName: str = Field(alias="output_zip_name")

class ExportResponse(BaseModel):
    model_config = ConfigDict(populate_by_name=True)
    success: bool
    filePath: str = Field(alias="file_path")
    warnings: List[str]

def ensure_temp_dir():
    os.makedirs(TEMP_DIR, exist_ok=True)

def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", name)

def write_cell(ws, cell: str, value: Any):
    ws[cell] = value

def fill_excel(template_path: str, sheet_mapping: SheetMapping, record_data: Dict[str, Any]) -> tuple[str, List[str]]:
    wb = load_workbook(template_path)
    ws = wb[sheet_mapping.sheet]
    warnings: List[str] = []
    for item in sheet_mapping.mappings:
        try:
            value = record_data.get(item.fieldId)
            if value is not None:
                write_cell(ws, item.cell, value)
        except Exception as e:
            warnings.append(f"Field {item.fieldId} cell {item.cell}: {str(e)}")
    file_id = str(uuid.uuid4())
    output_path = os.path.join(TEMP_DIR, f"{file_id}.xlsx")
    wb.save(output_path)
    return output_path, warnings

@app.get("/health")
def health_check():
    return {"status": "ok"}

@app.post("/export/single", response_model=ExportResponse)
def export_single(request: SingleExportRequest):
    if not os.path.isfile(request.templatePath):
        raise HTTPException(status_code=400, detail="Template file not found")
    ensure_temp_dir()
    try:
        sheet_mapping = SheetMapping(sheet=request.sheet, mappings=request.mappings)
        file_path, warnings = fill_excel(request.templatePath, sheet_mapping, request.recordData)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
    return ExportResponse(success=True, filePath=file_path, warnings=warnings)

@app.post("/export/batch", response_model=ExportResponse)
def export_batch(request: BatchExportRequest):
    if not os.path.isfile(request.templatePath):
        raise HTTPException(status_code=400, detail="Template file not found")
    if not request.records:
        raise HTTPException(status_code=400, detail="No records provided")
    ensure_temp_dir()
    warnings: List[str] = []
    temp_files: List[str] = []
    used_names: Dict[str, int] = {}
    for record in request.records:
        try:
            sheet_mapping = SheetMapping(sheet=request.sheet, mappings=request.mappings)
            file_path, file_warnings = fill_excel(request.templatePath, sheet_mapping, record.data)
            temp_files.append(file_path)
            warnings.extend(file_warnings)
        except Exception as e:
            warnings.append(f"Export failed for {record.outputName}: {str(e)}")
            continue
    if not temp_files:
        raise HTTPException(status_code=500, detail="All exports failed")
    zip_id = str(uuid.uuid4())
    zip_path = os.path.join(TEMP_DIR, f"{zip_id}.zip")
    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for idx, file_path in enumerate(temp_files):
                base_name = sanitize_filename(request.records[idx].outputName)
                if not base_name.lower().endswith(".xlsx"):
                    base_name += ".xlsx"
                count = used_names.get(base_name, 0)
                used_names[base_name] = count + 1
                if count > 0:
                    name, ext = os.path.splitext(base_name)
                    final_name = f"{name}({count}){ext}"
                else:
                    final_name = base_name
                zf.write(file_path, final_name)
    except Exception as e:
        for file_path in temp_files:
            if os.path.exists(file_path): os.remove(file_path)
        raise HTTPException(status_code=500, detail=f"ZIP creation failed: {str(e)}")
    for file_path in temp_files:
        if os.path.exists(file_path): os.remove(file_path)
    return ExportResponse(success=True, filePath=zip_path, warnings=warnings)
